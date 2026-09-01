#!/usr/bin/env python3
"""
Kandydaci na pytania do panelu - z DANYCH, nie z glowy.

Scala trzy zrodla z data/pytania/zrodla/ w klastry intencji, osobno dla kazdej strony:
  * Google Search Console  - gsc-zapytania.csv  (wyswietlenia, kliki, pozycja)
  * Bing Webmaster Tools   - bing-zapytania.csv (to samo + zapytania z AI)
  * Senuto                 - senuto-widocznosc.xlsx (SREDNI MIESIECZNY WOLUMEN)

Po co trzy: GSC i Bing pokazuja tylko to, na co JUZ jestesmy widoczni - zapytanie,
na ktore nie mamy tresci, nie pojawi sie tam wcale. Senuto podaje wolumen niezalezny
od naszej widocznosci, wiec dopiero razem widac i popyt, i luki.

UWAGA o eksporcie GSC: plik zapytan jest dla CALEJ domeny (GSC nie eksportuje przekroju
zapytanie x strona), wiec przypisanie do stron robimy po TEMACIE - wzorcami nizej.
Klaster "inne" jest po to, zeby bylo widac, czego wzorce nie zlapaly.

Uruchom:  python scripts/gen-kandydaci-pytan.py           (wszystkie strony)
          python scripts/gen-kandydaci-pytan.py prawo-mol (jedna)
Wynik:    data/pytania/kandydaci-<slug>.md
"""
import csv, io, os, re, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ZRODLA = os.path.join(ROOT, "data", "pytania", "zrodla")

# Dla kazdej strony: co uznajemy za jej temat, co odsiewamy jako obce, i jak dzielimy
# na klastry. Kolejnosc klastrow ma znaczenie - zapytanie trafia do PIERWSZEGO
# pasujacego, wiec waskie wzorce musza stac nad szerokimi.
STRONY = {
    "prawo-rhd": {
        "temat": (r"\brhd\b|rolnicz\w+ handel|rolnicza sprzeda|handel detaliczn|ewidencj\w+ sprzeda"
                  r"|sprzeda\w* (?:ser|przetwor|p\w+od)|rolnik mo\w+ sprzeda|sprzedawa\w+ w\w+asne"
                  r"|w\w+asnego gospodarstwa|sprzeda\w+ bezpo\w+redni"),
        # RHD to takze "right-hand drive" - auta z kierownica po prawej. Samo "rhd plus"
        # to 210 wyszukiwan miesiecznie ruchu, ktory nigdy nie szukal sera.
        "obce": r"samochod|\bauto|kierownic|nowogard|stacja|paliw|\brhd plus\b|\ba rhd \+",
        "klastry": [
            ("kto-moze", "Kto może prowadzić RHD — czy trzeba być rolnikiem?",
             r"kto mo\w+ prowadzi|czy trzeba by\w+ rolnikiem|rolnik mo\w+ sprzedawa|co mo\w+ sprzedawa\w+ rolnik"),
            ("co-mozna-sprzedawac", "Czy mogę sprzedawać sery własnej produkcji w ramach RHD?",
             r"sprzeda\w+ ser\w+ w\w+asnej|rhd co mo\w+na sprzedawa|sprzeda\w+ przetwor\w+|sprzeda\w+ w ramach rhd|sprzedawa\w+ w\w+asne przetwory|p\w+od\w+w rolnych z w\w+asnego"),
            ("ewidencja", "Co musi zawierać ewidencja sprzedaży w RHD?", r"ewidencj"),
            ("zgloszenie-us", "Czy RHD trzeba zgłaszać do urzędu skarbowego?",
             r"urz\w+du skarbowego|rhd.*podatek|handel detaliczny podatek|detaliczny a vat"),
            ("limit-przychodu", "Jaki jest limit przychodu w RHD?", r"limit"),
            ("kasa-faktura", "Czy w RHD potrzebna jest kasa fiskalna?", r"kasa fiskaln|faktur|paragon"),
            ("rejestracja", "Gdzie i kiedy zarejestrować działalność RHD?",
             r"rejestracj|sanepid|wniosek|jak zacz\w+|zg\w+oszenie|wymog\w* sanitarn|wymagani|wymogi|ustawa"),
            ("vs-sprzedaz-bezposrednia", "Czym różni się RHD od sprzedaży bezpośredniej?",
             r"sprzeda\w+ bezpo\w+redni|rolnicza sprzeda\w+ detaliczna|handel rolniczy|handel p\w+odami"),
            ("numer-rhd", "Co to jest numer RHD i jak wygląda?", r"numer rhd|wykaz"),
            ("przez-internet", "Czy w ramach RHD można sprzedawać przez internet?", r"przez internet"),
            ("co-to-rhd", "Co to jest RHD?", r"co to|co znaczy|znaczy rhd|^rhd$|^rolniczy handel detaliczny$"),
            ("miod-jaja", "(poza tematem serowarskim: miód, jaja)", r"mi\w+d|jaj"),
        ],
    },
    "prawo-mol": {
        "temat": (r"\bmol\b|\bmlo\b|marginaln\w+ (?:lokaln|ograniczon)|lokaln\w+ i ograniczon"
                  r"|ma\w+\w+ skal|masarni"),
        # MOL to takze wegierski koncern paliwowy - "stacja paliw mol", "paliwo mol".
        # Ta sama pulapka co RHD i right-hand drive.
        "obce": r"stacja|paliw|benzyn|w\w+giers|z jakiego kraju|\bmol\.|molo\b",
        "klastry": [
            ("mol-vs-rhd", "Czym MOL różni się od RHD — co wybrać?", r"\brhd\b"),
            ("limity", "Jakie są limity produkcji sera w MOL?", r"limit|ile mo\w+na|tygodniow"),
            ("rejestracja", "Jak zarejestrować działalność MOL?",
             r"rejestracj|zg\w+oszenie|wniosek|weterynari|sanepid|jak zacz|wymog|wymagani"),
            ("gdzie-sprzedawac", "Gdzie wolno sprzedawać produkty z MOL?",
             r"gdzie|obszar|sprzeda\w+ do|sklep|restauracj|powiat|wojew"),
            ("wlasny-surowiec", "Czy w MOL surowiec musi pochodzić z własnego gospodarstwa?",
             r"surowc|surowiec|w\w+asneg|gospodarstw"),
            ("co-mozna-produkowac", "Co można produkować w ramach MOL?",
             r"co mo\w+na|produkcj\w* (?:na|zywnosci)|ma\w+\w+ skal|masarni|zak\w+ad"),
            ("co-to-mol", "Co to jest MOL i co oznacza ten skrót?",
             r"co to|skr\w+t|znaczy|dzia\w+alno|marginaln|^mol$|\bmlo\b|definicj"),
        ],
    },
}


def bez_ogonkow(t):
    return t.translate(str.maketrans("ąćęłńóśźż", "acelnoszz"))


def licz(t):
    """Liczba z eksportu: Bing/Senuto uzywaja przecinka jako separatora dziesietnego."""
    if t is None:
        return 0.0
    t = str(t).strip().replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def czytaj_gsc():
    with io.open(os.path.join(ZRODLA, "gsc-zapytania.csv"), encoding="utf-8-sig", newline="") as f:
        for w in csv.DictReader(f):
            q = (w.get("Najczęstsze zapytania") or "").strip()
            if q:
                yield q, licz(w.get("Wyświetlenia")), licz(w.get("Kliknięcia")), licz(w.get("Pozycja"))


def czytaj_bing():
    with io.open(os.path.join(ZRODLA, "bing-zapytania.csv"), encoding="utf-8-sig", newline="") as f:
        for w in csv.DictReader(f):
            q = (w.get("Słowo kluczowe") or "").strip()
            if q:
                yield q, licz(w.get("Wyświetlenia")), licz(w.get("Kliknięcia")), licz(w.get("Śr. pozycja"))


def czytaj_senuto():
    try:
        import openpyxl
    except ImportError:
        print("  (openpyxl niedostepny - pomijam Senuto)", file=sys.stderr)
        return
    skoroszyt = openpyxl.load_workbook(os.path.join(ZRODLA, "senuto-widocznosc.xlsx"),
                                       read_only=True, data_only=True)
    arkusz = skoroszyt[skoroszyt.sheetnames[0]]
    for i, wiersz in enumerate(arkusz.iter_rows(values_only=True)):
        if i == 0 or not wiersz or not wiersz[0]:
            continue
        yield str(wiersz[0]).strip(), licz(wiersz[1]), licz(wiersz[2])


def przetworz(slug, opis):
    temat = re.compile(bez_ogonkow(opis["temat"]), re.I)
    obce = re.compile(bez_ogonkow(opis["obce"]), re.I)
    klastry = opis["klastry"]

    def dopasuj(q):
        plaskie = bez_ogonkow(q.lower())
        for klucz, _t, wzor in klastry:
            if re.search(bez_ogonkow(wzor), plaskie):
                return klucz
        return "inne"

    zebrane = {k: {"gsc": [], "bing": [], "senuto": []} for k, _, _ in klastry}
    zebrane["inne"] = {"gsc": [], "bing": [], "senuto": []}

    def pasuje(q):
        p = bez_ogonkow(q)
        return temat.search(p) and not obce.search(p)

    for q, wysw, kliki, poz in czytaj_gsc():
        if pasuje(q):
            zebrane[dopasuj(q)]["gsc"].append((q, wysw, kliki, poz))
    for q, wysw, kliki, poz in czytaj_bing():
        if pasuje(q):
            zebrane[dopasuj(q)]["bing"].append((q, wysw, kliki, poz))
    for q, wol, poz in czytaj_senuto():
        if pasuje(q):
            zebrane[dopasuj(q)]["senuto"].append((q, wol, poz))

    tytuly = dict((k, t) for k, t, _ in klastry)
    tytuly["inne"] = "(niesklasyfikowane — przejrzyj i dopisz wzorzec)"

    wiersze = []
    for klucz in tytuly:
        d = zebrane[klucz]
        wiersze.append({
            "klucz": klucz, "tytul": tytuly[klucz],
            "gsc_wysw": sum(x[1] for x in d["gsc"]),
            "gsc_kliki": sum(x[2] for x in d["gsc"]),
            "senuto_wolumen": sum(x[1] for x in d["senuto"]),
            "bing_wysw": sum(x[1] for x in d["bing"]),
            "zapytan": len(d["gsc"]) + len(d["bing"]) + len(d["senuto"]),
            "dane": d,
        })
    # Sortujemy po wolumenie Senuto, potem po wyswietleniach GSC: wolumen mowi o POPYCIE,
    # wyswietlenia tylko o naszej obecnej widocznosci.
    wiersze.sort(key=lambda r: (r["senuto_wolumen"], r["gsc_wysw"]), reverse=True)

    linie = ["# Kandydaci na pytania — /%s" % slug.replace("-", "/", 1), "",
             "PLIK GENEROWANY: `python scripts/gen-kandydaci-pytan.py`. Nie edytuj ręcznie —",
             "zmieniaj klastry w skrypcie albo eksporty w `data/pytania/zrodla/`.", "",
             "**Senuto** to średni miesięczny wolumen wyszukiwań (popyt niezależny od naszej",
             "widoczności). **GSC**/**Bing** to wyświetlenia, czyli gdzie już nas widać.", "",
             "| Klaster | Senuto/mies. | GSC wyśw. | GSC klik. | Bing wyśw. | zapytań |",
             "|---|---:|---:|---:|---:|---:|"]
    for r in wiersze:
        linie.append("| %s — %s | %d | %d | %d | %d | %d |" % (
            r["klucz"], r["tytul"], r["senuto_wolumen"], r["gsc_wysw"],
            r["gsc_kliki"], r["bing_wysw"], r["zapytan"]))

    linie += ["", "## Zapytania w klastrach", ""]
    for r in wiersze:
        if not r["zapytan"]:
            continue
        linie.append("### %s — %s" % (r["klucz"], r["tytul"]))
        if r["dane"]["senuto"]:
            linie.append("**Senuto (wolumen/mies., pozycja)**")
            for q in sorted(r["dane"]["senuto"], key=lambda x: -x[1]):
                linie.append("- `%s` — %d, poz. %s" % (q[0], q[1], q[2]))
        for nazwa, dane in (("GSC (wyświetlenia, kliknięcia, pozycja)", r["dane"]["gsc"]),
                            ("Bing (wyświetlenia, kliknięcia, pozycja)", r["dane"]["bing"])):
            if dane:
                linie.append("**%s**" % nazwa)
                for q in sorted(dane, key=lambda x: -x[1]):
                    linie.append("- `%s` — %d wyśw., %d klik., poz. %.1f" % (q[0], q[1], q[2], q[3]))
        linie.append("")

    wynik = os.path.join(ROOT, "data", "pytania", "kandydaci-%s.md" % slug)
    io.open(wynik, "w", encoding="utf-8", newline="\n").write("\n".join(linie) + "\n")

    print("== %s -> %s" % (slug, os.path.relpath(wynik, ROOT)))
    print("   %-26s %9s %9s %9s" % ("KLASTER", "SENUTO", "GSC", "BING"))
    for r in wiersze:
        if r["zapytan"]:
            print("   %-26s %9d %9d %9d" % (r["klucz"], r["senuto_wolumen"], r["gsc_wysw"], r["bing_wysw"]))
    print()


def main():
    wybrane = [a for a in sys.argv[1:] if not a.startswith("-")] or list(STRONY)
    for slug in wybrane:
        if slug not in STRONY:
            raise SystemExit("Nie znam strony '%s'. Dostepne: %s" % (slug, ", ".join(STRONY)))
        przetworz(slug, STRONY[slug])


if __name__ == "__main__":
    main()
